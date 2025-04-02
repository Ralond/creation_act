from datetime import datetime
from pathlib import Path
import openpyxl
import time
from copy import copy
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR, CERTIFICATE_EXTENSIONS

class FileManager:
    @staticmethod
    def safe_get(sequence, index, default=None):
        """Безопасное получение элемента"""
        try:
            if isinstance(sequence, (list, tuple)) and index < len(sequence):
                return sequence[index] if sequence[index] is not None else default
            return default
        except Exception:
            return default

    @staticmethod
    def load_workbook_safe(path, max_attempts=3):
        """Загрузка Excel файла с повторами"""
        for attempt in range(max_attempts):
            try:
                if not Path(path).exists():
                    raise FileNotFoundError(f"Файл не найден: {path}")
                return openpyxl.load_workbook(path, data_only=True)
            except Exception as e:
                print(f"Ошибка загрузки (попытка {attempt+1}): {e}")
                if attempt == max_attempts - 1:
                    raise
                time.sleep(1)

    @staticmethod
    def create_akt_folder(akt_num):
        """Создает папку для акта"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_name = f"Акт_{akt_num}_{timestamp}"
        folder_path = OUTPUT_DIR / folder_name
        folder_path.mkdir(parents=True, exist_ok=True)
        return folder_path

    @staticmethod
    def save_workbook_safe(wb, filepath, max_attempts=3):
        """Безопасное сохранение файла"""
        temp_path = filepath.with_name(f"temp_{filepath.name}")
        
        for attempt in range(max_attempts):
            try:
                wb.save(temp_path)
                if filepath.exists():
                    filepath.unlink()
                temp_path.rename(filepath)
                return True
            except Exception as e:
                print(f"Ошибка сохранения (попытка {attempt+1}): {e}")
                time.sleep(1)
        return False

    @staticmethod
    def validate_certificate(filepath):
        """Проверка формата сертификата"""
        return filepath.suffix.lower() in CERTIFICATE_EXTENSIONS