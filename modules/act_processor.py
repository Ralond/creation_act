import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path
from datetime import datetime
from modules.file_manager import FileManager

class ActProcessor:
    def __init__(self):
        self.file_manager = FileManager()
        self.organizations = {}
        self.personnel = {}
        self.normatives = {}
        self.certificates = {}

    def load_source_data(self, register_path):
        """Загрузка всех данных из реестра АОСР"""
        wb = self.file_manager.load_workbook_safe(register_path)
        
        # Загрузка организаций
        org_sheet = wb['Организации']
        for row in org_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Проверяем, что есть ID организации
                self.organizations[row[0]] = {
                    'type': row[1],
                    'name': row[2],
                    'ogrn': row[3],
                    'inn': row[4],
                    'address': row[5],
                    'phone': row[6],
                    'sro': row[7]
                }
        
        # Загрузка персоналий
        pers_sheet = wb['Персоналии']
        for row in pers_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Проверяем, что есть ID персоналии
                self.personnel[row[0]] = {
                    'name': row[1],
                    'position': row[2],
                    'organization': row[3],
                    'role': row[4],
                    'phone': row[5],
                    'nrs': row[6],
                    'order': row[7],
                    'active_from': row[8]
                }
        
        # Загрузка нормативов
        norms_sheet = wb['Нормативы']
        for row in norms_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Проверяем, что есть код норматива
                self.normatives[row[0]] = {
                    'name': row[1],
                    'type': row[2],
                    'full_name': row[3],
                    'status': row[4]
                }
        
        # Загрузка сертификатов
        cert_sheet = wb['Сертификаты']
        for row in cert_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Проверяем, что есть ID сертификата
                self.certificates[row[0]] = {
                    'material': row[1],
                    'cert_number': row[2],
                    'full_info': row[3],
                    'manufacturer': row[4]
                }
        
        wb.close()

    def process_register(self, register_path):
        """Обработка реестра актов"""
        self.load_source_data(register_path)
        
        wb = self.file_manager.load_workbook_safe(register_path)
        sheet = wb['Реестр актов']
        
        valid_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if self.validate_row(row):
                valid_rows.append(row)
            elif any(cell is not None for cell in row):
                break  # Прекращаем при первой частично заполненной строке
                
        wb.close()
        return valid_rows

    def validate_row(self, row):
        """Проверка валидности строки реестра"""
        if not isinstance(row, (list, tuple)):
            return False
        
        if all(cell is None for cell in row):
            return False
        
        if len(row) < 7:
            return False
        
        if not (row[0] and row[5] and isinstance(row[5], datetime)):
            return False
        
        return True

    def generate_all_akts(self, rows, template_path, output_path):
        """Генерация всех актов в одной книге"""
        try:
            # Загружаем шаблон
            wb_template = self.file_manager.load_workbook_safe(template_path)
            template_sheet = wb_template.active
            
            # Создаем новую книгу для всех актов
            output_wb = openpyxl.Workbook()
            output_wb.remove(output_wb.active)  # Удаляем дефолтный лист
            
            success_count = 0
            for row in rows:
                akt_id = f"{row[0]}-{row[1]}"
                
                try:
                    # Создаем новый лист для акта
                    sheet_name = f"Акт {akt_id}"[:31]  # Ограничение длины имени листа
                    new_sheet = output_wb.create_sheet(title=sheet_name)
                    
                    # Копируем шаблон
                    self._copy_template(template_sheet, new_sheet)
                    
                    # Заполняем данные
                    self._fill_akt_data(new_sheet, row)
                    
                    success_count += 1
                    
                except Exception as e:
                    print(f"Ошибка при обработке акта {akt_id}: {str(e)}")
                    continue
            
            # Сохраняем файл со всеми актами
            self.file_manager.save_workbook_safe(output_wb, output_path)
            
            return {
                'file': output_path,
                'total': len(rows),
                'success': success_count,
                'status': 'success'
            }
            
        except Exception as e:
            return {
                'file': output_path,
                'error': str(e),
                'status': 'error'
            }
        finally:
            if 'wb_template' in locals():
                wb_template.close()
            if 'output_wb' in locals():
                output_wb.close()

    def _copy_template(self, template_sheet, new_sheet):
        """Копирование шаблона с сохранением стилей"""
        # Копируем объединенные ячейки
        for merged_range in template_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))
        
        # Копируем размеры столбцов
        for col in template_sheet.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter in template_sheet.column_dimensions:
                new_sheet.column_dimensions[col_letter].width = template_sheet.column_dimensions[col_letter].width
        
        # Копируем высоты строк
        for row_idx, row_dim in template_sheet.row_dimensions.items():
            new_sheet.row_dimensions[row_idx].height = row_dim.height
        
        # Копируем ячейки с сохранением стилей
        for row in template_sheet.iter_rows():
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                    
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value
                
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

    def _fill_akt_data(self, sheet, row):
        """Заполнение данных акта в шаблоне"""
        # Основные данные акта
        self._write_to_cell(sheet, 'C8', f"{row[0]}-{row[1]}")  # Номер акта
        self._write_to_cell(sheet, 'K8', row[5].day)  # День акта
        self._write_to_cell(sheet, 'O8', row[5].month)  # Месяц акта
        self._write_to_cell(sheet, 'S8', row[5].year)  # Год акта
        
        # Наименование работ
        self._write_to_cell(sheet, 'A45', row[3])  # Наименование работ
        
        # Даты выполнения работ
        start_date = row[4]  # Дата начала
        end_date = row[5]  # Дата окончания
        self._write_to_cell(sheet, 'K56', start_date.day if start_date else "")
        self._write_to_cell(sheet, 'O56', start_date.month if start_date else "")
        self._write_to_cell(sheet, 'S56', start_date.year if start_date else "")
        self._write_to_cell(sheet, 'K57', end_date.day if end_date else "")
        self._write_to_cell(sheet, 'O57', end_date.month if end_date else "")
        self._write_to_cell(sheet, 'S57', end_date.year if end_date else "")
        
        # Проектная документация
        self._write_to_cell(sheet, 'A49', f"Проект: {row[11]}, Лист: {row[12]}")  # Проект и лист
        
        # Материалы
        materials = []
        if row[8]:  # Материалы
            materials.append(row[8])
        if row[9]:  # Материалы вручную
            materials.append(row[9])
        self._write_to_cell(sheet, 'A51', "\n".join(materials))
        
        # Исполнительные схемы
        if row[10]:  # Исп. схемы
            self._write_to_cell(sheet, 'F59', row[10])
        
        # Нормативные документы
        if row[13]:  # Нормативные документы
            self._write_to_cell(sheet, 'A61', row[13])
        
        # Последующие работы
        if row[7]:  # Последующие работы
            self._write_to_cell(sheet, 'A63', row[7])
        
        # Заполнение информации об организациях и персонах
        self._fill_organization_data(sheet)
        
        # Примечания
        if row[14]:  # Примечания
            self._write_to_cell(sheet, 'K65', row[14])
        
        # Приложения
        attachments = []
        if row[10]:  # Исп. схемы
            attachments.append(f"Исполнительные схемы: {row[10]}")
        if row[8] or row[9]:  # Материалы
            attachments.append("Сертификаты на материалы")
        self._write_to_cell(sheet, 'A68', "\n".join(attachments))

    def _fill_organization_data(self, sheet):
        """Заполнение данных об организациях и персонах"""
        # Пример заполнения данных заказчика
        customer_org = next((org for org in self.organizations.values() if org['type'] == 'Заказчик'), None)
        if customer_org:
            self._write_to_cell(sheet, 'A11', customer_org['name'])
            self._write_to_cell(sheet, 'A12', f"ОГРН: {customer_org['ogrn']}, ИНН: {customer_org['inn']}")
            self._write_to_cell(sheet, 'A13', f"Адрес: {customer_org['address']}, Тел.: {customer_org['phone']}")
        
        # Заполнение данных представителей
        customer_rep = next((p for p in self.personnel.values() if p['role'] == 'Заказчик'), None)
        if customer_rep:
            self._write_to_cell(sheet, 'A21', f"{customer_rep['position']}, {customer_rep['name']}")
            self._write_to_cell(sheet, 'A22', f"НРС: {customer_rep['nrs']}, Приказ: {customer_rep['order']}")
            self._write_to_cell(sheet, 'A72', customer_rep['name'])  # Подпись
            
        # Аналогично заполняем данные для других ролей

    def _write_to_cell(self, sheet, coord, value):
        """Безопасная запись в ячейку"""
        cell = sheet[coord]
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            cell.value = value
            return True
        return False