import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

def get_register_data(register_path, akt_id):
    """
    Получает данные акта из реестра по ID акта
    
    :param register_path: Путь к файлу реестра
    :param akt_id: ID акта для поиска (например, "АСР-2023-001")
    :return: Словарь с данными акта или None если не найден
    """
    try:
        wb = openpyxl.load_workbook(register_path)
        ws = wb["Реестр актов"]
        
        # Находим строку с нужным актом
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == akt_id:  # ID акта в первом столбце
                # Собираем данные из строки реестра
                data = {
                    'akt_number': f"{row[1]}-{row[2]}",  # Суффикс + Номер
                    'works': row[3],  # Наименование работ
                    'start_date': row[4],  # Дата начала
                    'end_date': row[5],  # Дата окончания
                    'akt_date': row[6],  # Дата акта
                    'next_works': row[7],  # Последующие работы
                    'materials': row[8],  # Материалы
                    'shemi': row[9],  # Исп. схемы
                    'project_docs': f"{row[10]}, лист {row[11]}",  # Проект + лист
                    'regulations': row[12],  # Нормативные документы
                    'notes': row[14],  # Примечания
                    'customer_rep': row[15],  # Предст. заказчика
                    'contractor_rep': row[16],  # Предст. генподрядчика
                    'tech_supervisor': row[17],  # Технадзор
                    'designer_rep': row[18],  # Проектировщик
                    'executor_rep': row[19]  # Исполнитель работ
                }
                return data
        return None
        
    except Exception as e:
        print(f"Ошибка при чтении реестра: {str(e)}")
        return None

def fill_aosr_template(template_path, output_path, register_data, persons_data, orgs_data):
    """
    Заполняет шаблон АОСР данными из реестра
    
    :param template_path: Путь к шаблону
    :param output_path: Путь для сохранения
    :param register_data: Данные акта из реестра
    :param persons_data: Данные персоналий
    :param orgs_data: Данные организаций
    :return: True если успешно, False если ошибка
    """
    try:
        # Загружаем шаблон
        wb = openpyxl.load_workbook(template_path)
        ws = wb.worksheets[0]
        
        # Парсим даты
        akt_date = datetime.strptime(register_data['akt_date'], "%d.%m.%Y")
        start_date = datetime.strptime(register_data['start_date'], "%d.%m.%Y")
        end_date = datetime.strptime(register_data['end_date'], "%d.%m.%Y")
        
        # Основные данные объекта (пример, нужно адаптировать под вашу структуру)
        project_name = "Объект капитального строительства"  # Должно браться из других данных
        ws['D6'] = project_name
        
        # Данные организаций (пример)
        customer_org = next((org for org in orgs_data if org['type'] == 'Заказчик'), None)
        if customer_org:
            ws['D9'] = f"{customer_org['name']}, ОГРН {customer_org['ogrn']}, ИНН {customer_org['inn']}, {customer_org['address']}"
        
        # Данные акта из реестра
        ws['C75'] = register_data['akt_number']
        ws['X75'] = akt_date.day
        ws['AB75'] = akt_date.month
        ws['AF75'] = akt_date.year
        
        # Работы и материалы
        ws['D87'] = register_data['works']
        ws['D91'] = register_data['project_docs']
        ws['P95'] = register_data['materials']
        ws['G101'] = register_data['shemi']
        
        # Даты работ
        ws['L107'] = start_date.day
        ws['P107'] = start_date.month
        ws['T107'] = start_date.year
        ws['L108'] = end_date.day
        ws['P108'] = end_date.month
        ws['T108'] = end_date.year
        
        # Нормативные документы и последующие работы
        ws['D112'] = register_data['regulations']
        ws['D118'] = register_data['next_works']
        ws['K122'] = register_data['notes']
        ws['H124'] = "3"  # Количество экземпляров
        
        # Подписи (пример, нужно адаптировать под вашу структуру)
        customer_person = next((p for p in persons_data if p['id'] == register_data['customer_rep']), None)
        if customer_person:
            ws['D137'] = customer_person['fio']
        
        # Сохраняем заполненный файл
        wb.save(output_path)
        print(f"Акт успешно создан: {output_path}")
        return True
    
    except Exception as e:
        print(f"Ошибка при заполнении шаблона: {str(e)}")
        return False

def get_persons_data(register_path):
    """Получает данные персоналий из реестра"""
    try:
        wb = openpyxl.load_workbook(register_path)
        ws = wb["Персоналии"]
        
        persons = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Если есть ID
                persons.append({
                    'id': row[0],
                    'fio': row[1],
                    'position': row[2],
                    'org': row[3],
                    'role': row[4],
                    'phone': row[5],
                    'nrs': row[6],
                    'order': row[7],
                    'valid_from': row[8]
                })
        return persons
        
    except Exception as e:
        print(f"Ошибка при чтении персоналий: {str(e)}")
        return []

def get_orgs_data(register_path):
    """Получает данные организаций из реестра"""
    try:
        wb = openpyxl.load_workbook(register_path)
        ws = wb["Организации"]
        
        orgs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Если есть ID
                orgs.append({
                    'id': row[0],
                    'type': row[1],
                    'name': row[2],
                    'ogrn': row[3],
                    'inn': row[4],
                    'address': row[5],
                    'phone': row[6],
                    'sro': row[7]
                })
        return orgs
        
    except Exception as e:
        print(f"Ошибка при чтении организаций: {str(e)}")
        return []

if __name__ == "__main__":
    # Пути к файлам
    register_path = "Реестр_АОСР_финальный.xlsx"
    template_path = "aosr-prikaz-344.xls"
    output_path = "АОСР_заполненный.xls"
    
    # ID акта для заполнения (можно ввести с клавиатуры)
    akt_id = "АСР-2023-001"
    
    # Получаем данные из реестра
    register_data = get_register_data(register_path, akt_id)
    if not register_data:
        print(f"Акт с ID {akt_id} не найден в реестре!")
        exit()
    
    persons_data = get_persons_data(register_path)
    orgs_data = get_orgs_data(register_path)
    
    # Заполняем шаблон
    fill_aosr_template(
        template_path=template_path,
        output_path=output_path,
        register_data=register_data,
        persons_data=persons_data,
        orgs_data=orgs_data
    )
    """
Ключевые особенности обновленного скрипта:
Интеграция с реестром:
get_register_data() - извлекает данные конкретного акта по его ID
get_persons_data() и get_orgs_data() - получают информацию о персоналиях и организациях

Автоматическое заполнение:
Основные данные акта берутся из строки реестра
Данные организаций и подписей подтягиваются из соответствующих листов

Обработка данных:
Автоматическое преобразование дат в нужный формат
Формирование комплексных полей (например, проект + лист)

Гибкость:
Можно легко адаптировать под конкретную структуру реестра
Поддерживается добавление новых полей

Обработка ошибок:
Проверка наличия акта в реестре
Защита от ошибок при чтении файлов

Для использования:
Убедитесь, что реестр имеет структуру как в предыдущем задании
Укажите правильные пути к файлам
Задайте ID акта, который нужно заполнить
Запустите скрипт - результат сохранится в указанный файл
Скрипт можно доработать для автоматического определения ID акта или пакетной обработки нескольких актов.
    """